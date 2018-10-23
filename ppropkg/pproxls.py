# -*- coding: utf-8 -*-

import os

import openpyxl

PPRO_PATH = os.getenv("HOMEDRIVE", "None") + os.getenv("HOMEPATH", "None") + "/Dropbox/POG/"
PO_HORSELIST_NAME = "POG_HorseList.xlsx"


class POHorseList:

    def __init__(self):
        self.wbpath = (PPRO_PATH + PO_HORSELIST_NAME).replace("\\", "/")
        self.wb = openpyxl.load_workbook(self.wbpath)
        self.ws = self.wb["POHorseList"]

        self.COL_OWNER_GENDER_RANK, self.COL_HORSE_NAME, self.COL_NAME_ORIGIN, self.COL_NK_URL, self.COL_NK_URL_SP, \
            self.COL_IS_SEALED, self.COL_NK_ID, self.COL_OWNER, self.COL_GENDER, self.COL_NOM_RANK, self.COL_STATUS, \
            self.COL_STATUS_OLD = 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12

    @staticmethod
    def _get_new_status(jra_status, status_old):
        if jra_status == "放牧":
            return "放牧"
        elif jra_status == "非放牧":
            if status_old in ("未登録", "登録", "抹消", "地方", "海外"):
                return "登録"
            elif status_old in ("在厩", "放牧"):
                return "在厩"
        elif jra_status == "未登録":
            if status_old == "未登録":
                return "未登録"
            elif status_old in ("登録", "在厩", "放牧"):
                return "抹消"
            elif status_old in ("抹消", "地方", "海外"):
                return status_old

    def update_status(self, status_list):
        is_updated = False

        for status_row in status_list:
            xlrow, jra_status = status_row[0], status_row[1]
            status_old = self.ws.cell(row=xlrow, column=self.COL_STATUS).value
            status_new = self._get_new_status(jra_status, status_old)

            if status_new != status_old:
                is_updated = True
                self.ws.cell(row=xlrow, column=self.COL_STATUS).value = status_new
            self.ws.cell(row=xlrow, column=self.COL_STATUS_OLD).value = status_old

        return is_updated

    def get(self):
        xlrow = 2
        mylist = []

        while self.ws.cell(row=xlrow, column=1).value:
            horse_name = self.ws.cell(row=xlrow, column=self.COL_HORSE_NAME).value
            owner_name = self.ws.cell(row=xlrow, column=self.COL_OWNER_GENDER_RANK).value
            horse_id = self.ws.cell(row=xlrow, column=self.COL_NK_ID).value
            is_seal = False if self.ws.cell(row=xlrow, column=self.COL_IS_SEALED).value == "-" else True
            mylist.append([horse_id, horse_name, owner_name, is_seal])
            xlrow += 1
        return mylist

    def get_name_list(self):
        xlrow = 2
        mylist = []
        while self.ws.cell(row=xlrow, column=self.COL_OWNER_GENDER_RANK).value:
            horse_name = self.ws.cell(row=xlrow, column=self.COL_HORSE_NAME).value
            if self.ws.cell(row=xlrow, column=self.COL_IS_SEALED).value == "-":
                mylist.append([horse_name,  xlrow])
            xlrow += 1
        return mylist

    def get_status_list(self, *args):
        xlrow = 2
        mylist = []

        while self.ws.cell(row=xlrow, column=self.COL_OWNER_GENDER_RANK).value:
            if self.ws.cell(row=xlrow, column=self.COL_IS_SEALED).value != "-":
                xlrow += 1
                continue
            status = self.ws.cell(row=xlrow, column=self.COL_STATUS).value
            status_old = self.ws.cell(row=xlrow, column=self.COL_STATUS_OLD).value
            if "updated_only" in args and status == status_old:
                xlrow += 1
                continue
            owner_gender_rank = self.ws.cell(row=xlrow, column=self.COL_OWNER_GENDER_RANK).value
            horse_name = self.ws.cell(row=xlrow, column=self.COL_HORSE_NAME).value
            horse_id = self.ws.cell(row=xlrow, column=self.COL_NK_ID).value
            owner = self.ws.cell(row=xlrow, column=self.COL_OWNER).value
            gender = self.ws.cell(row=xlrow, column=self.COL_GENDER).value
            nom_rank = self.ws.cell(row=xlrow, column=self.COL_NOM_RANK).value
            if "updated_only" in args:
                mylist.append([horse_name, owner_gender_rank, horse_id, status, status_old])
            else:
                mylist.append([owner, gender, nom_rank, horse_name, horse_id, status, status_old])

            xlrow += 1

        return mylist

    def save(self):
        self.wb.save(self.wbpath)

    def close(self):
        self.wb.close()
